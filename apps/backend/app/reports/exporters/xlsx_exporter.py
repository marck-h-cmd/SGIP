"""XLSX Exporter using openpyxl - Professional layout in Spanish"""
from app.reports.base import BaseExporter, ReportData
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io


# Spanish labels
SUMMARY_LABELS = {
    "total_readings": "Lecturas Totales",
    "avg_pressure": "Presion Promedio",
    "avg_flow": "Caudal Promedio",
    "nrw_percentage": "Perdidas No Renumeradas",
    "water_loss_m3": "Perdidas de Agua",
    "water_loss_estimate": "Perdidas de Agua",
    "anomalies_total": "Anomalias Totales",
    "anomalies_critical": "Anomalias Criticas",
    "anomalies_detected": "Anomalias Detectadas",
    "incidents_total": "Incidencias Totales",
    "incidents_open": "Incidencias Abiertas",
    "incidents_created": "Incidencias Creadas",
    "incidents_resolved": "Incidencias Resueltas",
    "total_anomalies": "Total Anomalias",
    "total_incidents": "Total Incidencias",
}

SUMMARY_UNITS = {
    "avg_pressure": "mca",
    "avg_flow": "LPS",
    "nrw_percentage": "%",
    "water_loss_m3": "m\u00b3/dia",
    "water_loss_estimate": "m\u00b3/dia",
}

# Styles
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SECTION_FONT = Font(bold=True, size=11, color="1F4E79")
LABEL_FONT = Font(bold=True, size=10)
VALUE_FONT = Font(size=10)
ALT_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


class XLSXExporter(BaseExporter):
    """Export reports to Excel format with professional styling"""

    def get_format(self) -> str:
        return "xlsx"

    def export(self, report_data: ReportData) -> bytes:
        wb = Workbook()

        # Summary sheet
        self._create_summary_sheet(wb.active, report_data)

        # Details sheet(s) based on report type
        if report_data.report_type == "daily":
            self._create_daily_detail_sheet(wb, report_data)
        elif report_data.report_type == "weekly":
            self._create_weekly_detail_sheet(wb, report_data)
        elif report_data.report_type == "custom":
            self._create_custom_detail_sheet(wb, report_data)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    # ------------------------------------------------------------------
    # Summary sheet
    # ------------------------------------------------------------------
    def _create_summary_sheet(self, ws, rd: ReportData):
        ws.title = "Resumen"
        ws.sheet_properties.tabColor = "1F4E79"

        # Title
        ws.merge_cells("A1:D1")
        title_map = {
            "daily": "Reporte Diario de Operaciones",
            "weekly": "Reporte Semanal de Operaciones",
            "custom": "Reporte Personalizado",
            "sla": "Reporte de Cumplimiento SLA",
        }
        ws["A1"] = title_map.get(rd.report_type, f"Reporte {rd.report_type.title()}")
        ws["A1"].font = Font(bold=True, size=16, color="1F4E79")
        ws["A1"].alignment = LEFT
        ws.row_dimensions[1].height = 28

        # Metadata
        meta = [
            ("Sector:", f"{rd.dma_name} ({rd.dma_id})"),
            ("Periodo:", f"{rd.period_start.strftime('%d/%m/%Y')} al {rd.period_end.strftime('%d/%m/%Y')}"),
            ("Generado:", rd.generated_at.strftime("%d/%m/%Y %H:%M")),
        ]
        row = 3
        for label, value in meta:
            ws.cell(row=row, column=1, value=label).font = LABEL_FONT
            ws.cell(row=row, column=2, value=value).font = VALUE_FONT
            row += 1

        # Summary section header
        row += 1
        ws.merge_cells(f"A{row}:D{row}")
        ws.cell(row=row, column=1, value="Resumen Ejecutivo").font = SECTION_FONT
        ws.cell(row=row, column=1).fill = SECTION_FILL
        for c in range(1, 5):
            ws.cell(row=row, column=c).fill = SECTION_FILL
        row += 1

        # Summary table headers
        headers = ["Metrica", "Valor", "Unidad"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = CENTER
        row += 1

        # Summary data
        summary = rd.summary or {}
        alt = False
        for key, value in summary.items():
            label = SUMMARY_LABELS.get(key, key.replace("_", " ").title())
            unit = SUMMARY_UNITS.get(key, "")

            ws.cell(row=row, column=1, value=label).font = LABEL_FONT
            ws.cell(row=row, column=1).border = THIN_BORDER

            val_cell = ws.cell(row=row, column=2, value=value)
            val_cell.font = VALUE_FONT
            val_cell.border = THIN_BORDER
            val_cell.alignment = CENTER

            ws.cell(row=row, column=3, value=unit).font = VALUE_FONT
            ws.cell(row=row, column=3).border = THIN_BORDER
            ws.cell(row=row, column=3).alignment = CENTER

            if alt:
                for c in range(1, 4):
                    ws.cell(row=row, column=c).fill = ALT_FILL
            alt = not alt
            row += 1

        # Column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 20

    # ------------------------------------------------------------------
    # Daily detail sheet
    # ------------------------------------------------------------------
    def _create_daily_detail_sheet(self, wb, rd: ReportData):
        details = rd.details or {}
        ws = wb.create_sheet(title="Detalle Diario")
        ws.sheet_properties.tabColor = "4472C4"
        row = 1

        # --- Anomalies breakdown ---
        ab = details.get("anomalies_breakdown", {})
        if ab:
            row = self._section_row(ws, row, "Desglose de Anomalias")
            headers = ["Total", "Criticas", "Alta Severidad", "Severidad Media", "Baja Severidad"]
            values = [ab.get("total", 0), ab.get("critical", 0), ab.get("high", 0), ab.get("medium", 0), ab.get("low", 0)]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.border = THIN_BORDER
                cell.alignment = CENTER
            row += 1
            for col, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.font = VALUE_FONT
                cell.border = THIN_BORDER
                cell.alignment = CENTER
            row += 2

        # --- Incidents breakdown ---
        ib = details.get("incidents_breakdown", {})
        if ib:
            row = self._section_row(ws, row, "Desglose de Incidencias")
            headers = ["Total", "Abiertas", "Resueltas", "Criticas Abiertas"]
            values = [ib.get("total", 0), ib.get("open", 0), ib.get("resolved", 0), ib.get("critical_open", 0)]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.border = THIN_BORDER
                cell.alignment = CENTER
            row += 1
            for col, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.font = VALUE_FONT
                cell.border = THIN_BORDER
                cell.alignment = CENTER
            row += 2

        # --- Pressure & Flow stats ---
        ps = details.get("pressure_stats", {})
        fs = details.get("flow_stats", {})
        if ps or fs:
            row = self._section_row(ws, row, "Estadisticas de Presion y Caudal")
            headers = ["Metrica", "Minimo", "Maximo"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.border = THIN_BORDER
                cell.alignment = CENTER
            row += 1
            if ps:
                for col, v in enumerate(["Presion (mca)", ps.get("min", 0), ps.get("max", 0)], 1):
                    cell = ws.cell(row=row, column=col, value=v)
                    cell.font = VALUE_FONT
                    cell.border = THIN_BORDER
                    cell.alignment = CENTER
                row += 1
            if fs:
                for col, v in enumerate(["Caudal (LPS)", fs.get("min", 0), fs.get("max", 0)], 1):
                    cell = ws.cell(row=row, column=col, value=v)
                    cell.font = VALUE_FONT
                    cell.border = THIN_BORDER
                    cell.alignment = CENTER
                row += 1
            row += 1

        # --- Readings table ---
        readings = details.get("readings", [])
        if readings:
            row = self._section_row(ws, row, f"Lecturas ({len(readings)} registros)")
            headers = ["Hora", "Presion (mca)", "Caudal (LPS)", "Anomalia"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.border = THIN_BORDER
                cell.alignment = CENTER
            row += 1
            alt = False
            for r in readings:
                ts = r.get("timestamp", "")
                if "T" in ts:
                    ts = ts.split("T")[1][:5]
                vals = [ts, r.get("pressure_mca", 0), r.get("flow_lps", 0), "Si" if r.get("is_anomaly") else "No"]
                for col, v in enumerate(vals, 1):
                    cell = ws.cell(row=row, column=col, value=v)
                    cell.font = VALUE_FONT
                    cell.border = THIN_BORDER
                    cell.alignment = CENTER
                    if alt:
                        cell.fill = ALT_FILL
                alt = not alt
                row += 1

        # Auto-width columns
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 18

    # ------------------------------------------------------------------
    # Weekly detail sheet
    # ------------------------------------------------------------------
    def _create_weekly_detail_sheet(self, wb, rd: ReportData):
        details = rd.details or {}
        ws = wb.create_sheet(title="Detalle Semanal")
        ws.sheet_properties.tabColor = "70AD47"
        row = 1

        # Daily stats table
        daily_stats = details.get("daily_stats", [])
        if daily_stats:
            row = self._section_row(ws, row, "Estadisticas Diarias")
            headers = ["Fecha", "Presion Prom (mca)", "Caudal Prom (LPS)", "Lecturas", "Anomalias", "Perdidas (m\u00b3)"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.border = THIN_BORDER
                cell.alignment = CENTER
            row += 1
            alt = False
            for s in daily_stats:
                vals = [
                    s.get("date", ""),
                    s.get("avg_pressure", 0),
                    s.get("avg_flow", 0),
                    s.get("readings_count", 0),
                    s.get("anomalies_count", 0),
                    round(s.get("water_loss", 0), 2),
                ]
                for col, v in enumerate(vals, 1):
                    cell = ws.cell(row=row, column=col, value=v)
                    cell.font = VALUE_FONT
                    cell.border = THIN_BORDER
                    cell.alignment = CENTER
                    if alt:
                        cell.fill = ALT_FILL
                alt = not alt
                row += 1
            row += 1

        # Trends
        at = details.get("anomaly_trend", {})
        it = details.get("incident_trend", {})
        if at or it:
            row = self._section_row(ws, row, "Tendencias")
            headers = ["Metrica", "Tendencia", "Cambio (%)"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.border = THIN_BORDER
                cell.alignment = CENTER
            row += 1
            trend_labels = {"RISING": "En Aumento", "FALLING": "En Disminucion", "STABLE": "Estable"}
            if at:
                for col, v in enumerate(["Anomalias", trend_labels.get(at.get("trend", ""), at.get("trend", "")), at.get("change_percentage", 0)], 1):
                    cell = ws.cell(row=row, column=col, value=v)
                    cell.font = VALUE_FONT
                    cell.border = THIN_BORDER
                    cell.alignment = CENTER
                row += 1
            if it:
                for col, v in enumerate(["Incidencias", trend_labels.get(it.get("trend", ""), it.get("trend", "")), it.get("change_percentage", 0)], 1):
                    cell = ws.cell(row=row, column=col, value=v)
                    cell.font = VALUE_FONT
                    cell.border = THIN_BORDER
                    cell.alignment = CENTER
                row += 1

        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 22

    # ------------------------------------------------------------------
    # Custom detail sheet
    # ------------------------------------------------------------------
    def _create_custom_detail_sheet(self, wb, rd: ReportData):
        details = rd.details or {}
        ws = wb.create_sheet(title="Detalle")
        ws.sheet_properties.tabColor = "ED7D31"
        row = 1

        # Anomalies
        anomalies = details.get("anomalies_list", [])
        if anomalies:
            row = self._section_row(ws, row, f"Anomalias Detectadas ({len(anomalies)})")
            headers = ["ID", "Fecha", "Severidad", "Estado"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.border = THIN_BORDER
                cell.alignment = CENTER
            row += 1
            alt = False
            for a in anomalies:
                vals = [a.get("id", ""), a.get("date", ""), a.get("severity", ""), a.get("status", "")]
                for col, v in enumerate(vals, 1):
                    cell = ws.cell(row=row, column=col, value=v)
                    cell.font = VALUE_FONT
                    cell.border = THIN_BORDER
                    cell.alignment = CENTER
                    if alt:
                        cell.fill = ALT_FILL
                alt = not alt
                row += 1
            row += 1

        # Incidents
        incidents = details.get("incidents_list", [])
        if incidents:
            row = self._section_row(ws, row, f"Incidencias ({len(incidents)})")
            headers = ["Codigo", "Titulo", "Prioridad", "Estado", "Fecha"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.border = THIN_BORDER
                cell.alignment = CENTER
            row += 1
            alt = False
            for inc in incidents:
                vals = [inc.get("code", ""), inc.get("title", ""), inc.get("priority", ""), inc.get("status", ""), inc.get("date", "")]
                for col, v in enumerate(vals, 1):
                    cell = ws.cell(row=row, column=col, value=v)
                    cell.font = VALUE_FONT
                    cell.border = THIN_BORDER
                    cell.alignment = CENTER
                    if alt:
                        cell.fill = ALT_FILL
                alt = not alt
                row += 1

        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 22

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------
    def _section_row(self, ws, row: int, title: str) -> int:
        ws.merge_cells(f"A{row}:F{row}")
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        for c in range(1, 7):
            ws.cell(row=row, column=c).fill = SECTION_FILL
        return row + 2
